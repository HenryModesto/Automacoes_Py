import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

class RelatorioModel:
    def __init__(self):
        self.colunas = [
            'Tarefa', 
            'Tipo', 
            'Dado',
            'Status',
            'Tempo Execução (s)',
            'Timestamp'
        ]
        
        self.estilo_cabecalho = Font(bold=True, color="FFFFFF")
        self.preenchimento_cabecalho = PatternFill(
            start_color="4F81BD", 
            end_color="4F81BD", 
            fill_type="solid"
        )
        self.alinhamento_central = Alignment(horizontal='center')
        self.estilo_sucesso = Font(color="00B050")
        self.estilo_falha = Font(color="FF0000")

    def criar_relatorio(self):
        """Cria um novo DataFrame para o relatório"""
        return pd.DataFrame(columns=self.colunas)

    def adicionar_tarefa_relatorio(self, df_relatorio, tarefa, tipo, dado, status, tempo_execucao):
        """Adiciona uma tarefa ao relatório"""
        try:
            nova_linha = {
                'Tarefa': str(tarefa),
                'Tipo': str(tipo),
                'Dado': str(dado),
                'Status': str(status),
                'Tempo Execução (s)': round(float(tempo_execucao), 2),
                'Timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            
            return pd.concat([df_relatorio, pd.DataFrame([nova_linha])], ignore_index=True)
        except Exception as e:
            print(f"Erro ao adicionar tarefa: {str(e)}")
            return df_relatorio

    def salvar_relatorio(self, df_relatorio, caminho_arquivo):
        """Salva o relatório em Excel com formatação"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Relatório de Execução"
            
            for col_num, col_name in enumerate(self.colunas, 1):
                cell = ws.cell(row=1, column=col_num, value=col_name)
                cell.font = self.estilo_cabecalho
                cell.fill = self.preenchimento_cabecalho
                cell.alignment = self.alinhamento_central
            
        
            for _, row in df_relatorio.iterrows():
                ws.append(row.tolist())
            
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=4):
                for cell in row:
                    if cell.value == "Sucesso":
                        cell.font = self.estilo_sucesso
                    elif cell.value == "Falha":
                        cell.font = self.estilo_falha
            
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 20
            
            wb.save(caminho_arquivo)
            return True
            
        except Exception as e:
            print(f"Erro ao salvar relatório: {str(e)}")
            return False